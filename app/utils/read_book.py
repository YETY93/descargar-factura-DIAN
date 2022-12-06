import sys
import openpyxl
import time
# import webdriver
from selenium import webdriver

# import get book
from app.utils import get_book
#import route driver
from app.const import const_project, const_bd
# import update column BD
from app.connections import update_fact

def read_book_path():
    list_doucments: list = get_book.list_dir_path()
    path_book: str = (list_doucments[0])
    wb = openpyxl.load_workbook(path_book)
    list_sheets: list = wb.sheetnames
    sheet_active = wb[list_sheets[0]]

    COLUM_CUFE_DOC: str = "AT"
    COLUM_ESTATE: str = "BG"
    COLUM_ID: str = "A"
    COLUM_ESTATE_UPDATE: str = "BH"

    print("#######################################")
    print("Elija una opción")
    print("1. Actulizar Datos BD")
    print("2. Descaragra XMl")
    print("O cualuqier tecla para finalizar")
    option= input(" 1 o 2 : ")
    print("#######################################")

    if option.strip() == "1":
        consecutive: int = 2
        column_name: str = "fac_path_file"
        for row in sheet_active.iter_rows():

            cufe:str = COLUM_CUFE_DOC + str(consecutive)
            id: str = COLUM_ID + str(consecutive)
            cell_cufe = sheet_active[cufe].value
            cell_id = str(sheet_active[id].value)
            cell_update = COLUM_ESTATE_UPDATE + str(consecutive)

            try:
                # This function update the column fac_efactura
                # with params
                # $table_name $column_name $value_update $id
                update_fact.update_colum_fac(
                    const_bd.TABLE_NAME, column_name, cell_cufe ,cell_id
                )

                sheet_active[cell_update].value = const_project.UPDATE_SUCCESFULL
                wb.save(path_book)  # Save the changes
            except Exception as e:
                print("Error en leer excel para actualizar column")
                print(e)

            consecutive += 1

    elif option.strip() == "2":
        try:
            consecutive: int = 2
            # start instance of navigator
            navigator = webdriver.Chrome(executable_path=const_project.PATH_DRIVER)
            # redirecte to PATH_LOGIN_DIAN
            navigator.get(const_project.PATH_LOGIN_DIAN)
            # wait 20 secons for configurate path download
            time.sleep(5)

            print("### Configure la ruta de descarga en la nueva ventana ###")
            accept_download = input("Presione cualquier tecla para continuar: ")

            for row in sheet_active.iter_rows():
                print("Se inicia la descarga de archivos")

                cell_activate_id_doc: str = COLUM_CUFE_DOC + str(consecutive) # Cell for id documento electronic
                cell_id_doc_value = sheet_active[cell_activate_id_doc].value
                link_download = const_project.PATH_DOWNLOAD_ZIP

                navigator.execute_script('''window.open(" ''' + link_download + cell_id_doc_value + ''' ","_blank");''')
                time.sleep(5)

                cell_download: str = COLUM_ESTATE + str(consecutive) # Cell for confirmed download
                sheet_active[cell_download].value = const_project.DOWNLOAD_ZIP_SUCCESFULL
                consecutive += 1

                wb.save(path_book) # Save the changes

        except Exception as e:
            print("posible en : read book")
            print(e)
        print("Ha finalizado la descarga de archivos")

    else:
        sys.exit(0)
